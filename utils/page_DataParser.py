import openpyxl

class ExcelUtil:
    def __init__(self, file_path, sheet_name):
        self.data = openpyxl.load_workbook(file_path)
        self.table = self.data[sheet_name]
        self.keys = [cell.value for cell in self.table[1]]
        self.row_num = self.table.max_row
        self.col_num = self.table.max_column

    def return_dict_data(self):
        if self.row_num <= 1:
            print("总行数小于1")
        else:
            r = []
            for row in self.table.iter_rows(min_row=2, max_row=self.row_num, values_only=True):
                s = dict(zip(self.keys, row))
                r.append(s)
            return r

import yaml
#带测试用例编号key一起读取！
def yaml_data_with_file(file_name, key):#file_name:yaml文件路径,key:yaml文件的key
    with open(file_name, "r") as f:
        data = yaml.load(f,Loader=yaml.FullLoader)[key]
        ret = list()
        for k in data.keys():
            v=data[k];
            v["key"]=k;
            ret.append(v);
        return ret

def readTxtOrCsv(filePath):#filePath:txt或csv的文件路径
    temp=[]
    fp=open(filePath,mode="r",encoding="utf-8")
    for line in fp:
        line=line.strip()
        temp.append(line)
    fp.close()

    headers=temp[0].split(",")
    values=[]
    for i in range(len(temp)):
        if i>=1:
            values.append(temp[i])

    ret=[]
    for x in values:
        data={}
        temp1=x.split(",")
        for y in range(len(headers)):
            k=headers[y]
            v=temp1[y]
            data[k]=v
        ret.append(data)
    return ret

import pymysql
def readMysql(host,user,pwd,dbname,sql):
    conn=pymysql.connect(host=host,user=user,password=pwd,database=dbname,charset='utf8')
    cursor=conn.cursor(pymysql.cursors.DictCursor)
    rows=None
    try:
        #执行SQL语句
        cursor.execute(sql)
        rows=cursor.fetchall() #[{},{}]

    except:
        rows=[]
        print(cursor._last_executed)

    #关闭数据库连接
    conn.close()
    return rows

#测试
# sql='select Sno,Sname,Ssex,Class from student'
# cesi_data_old=readMysql("localhost","root","123456","test",sql)
# print(cesi_data_old)
#带日期示例
    # # # 情况7,从mysql取数据
    # sql='select Sno,Sname,Ssex,Sbirthday,Class from student'
    # @pytest.mark.parametrize("data",readMysql('localhost','root','123456','test',sql))
    # def test_007(self,data):
    #     brithday = data["Sbirthday"]
    #     # 把日期类型的brithday转换为字符串
    #     brithday_str = brithday.strftime("%Y-%m-%d")
    #     id = data["Sno"]
    #     name = data["Sname"]
    #     sex = data["Ssex"]
    #     class1 = data["Class"]
    #     birthday = brithday_str
    #     print(id, name, sex, class1, birthday)

